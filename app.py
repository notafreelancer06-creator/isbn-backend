from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
import time

app = Flask(__name__)
CORS(app)  # This allows your Vercel site to call this server

# ── helpers ────────────────────────────────────────────────────────────────

def fetch_google_books(isbn):
    try:
        url = f"https://www.googleapis.com/books/v1/volumes?q=isbn:{isbn}"
        r = requests.get(url, timeout=8)
        data = r.json()
        if data.get("totalItems", 0) == 0:
            return {}
        item = data["items"][0]["volumeInfo"]
        return {
            "title":     item.get("title", ""),
            "authors":   ", ".join(item.get("authors", [])),
            "publisher": item.get("publisher", ""),
            "year":      (item.get("publishedDate", "") or "")[:4],
            "pages":     item.get("pageCount", ""),
            "language":  item.get("language", ""),
            "cover":     (item.get("imageLinks", {}) or {}).get("thumbnail", ""),
            "description": item.get("description", "")[:300] if item.get("description") else "",
        }
    except Exception:
        return {}

def fetch_open_library(isbn):
    try:
        url = f"https://openlibrary.org/api/books?bibkeys=ISBN:{isbn}&format=json&jscmd=data"
        r = requests.get(url, timeout=8)
        data = r.json()
        key = f"ISBN:{isbn}"
        if key not in data:
            return {}
        item = data[key]
        authors = ", ".join(a.get("name", "") for a in item.get("authors", []))
        publishers = ", ".join(p.get("name", "") for p in item.get("publishers", []))
        return {
            "title":     item.get("title", ""),
            "authors":   authors,
            "publisher": publishers,
            "year":      (item.get("publish_date", "") or "")[-4:],
            "pages":     item.get("number_of_pages", ""),
            "cover":     item.get("cover", {}).get("medium", ""),
            "subjects":  ", ".join(s.get("name", "") for s in item.get("subjects", [])[:5]),
        }
    except Exception:
        return {}

def fetch_amazon(isbn):
    """Scrape Amazon India search results for the ISBN."""
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept-Language": "en-IN,en;q=0.9",
        }
        url = f"https://www.amazon.in/s?k={isbn}"
        r = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(r.text, "lxml")

        # Try to find price
        price = ""
        price_el = soup.select_one(".a-price .a-offscreen")
        if price_el:
            price = price_el.get_text(strip=True)

        # Try to find title
        title = ""
        title_el = soup.select_one("h2 a span")
        if title_el:
            title = title_el.get_text(strip=True)

        # Build direct search link
        link = f"https://www.amazon.in/s?k={isbn}"

        return {"title": title, "price": price, "link": link}
    except Exception:
        return {"link": f"https://www.amazon.in/s?k={isbn}"}

def fetch_flipkart(isbn):
    """Scrape Flipkart search results for the ISBN."""
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        }
        url = f"https://www.flipkart.com/search?q={isbn}&otracker=search"
        r = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(r.text, "lxml")

        # Try different selectors (Flipkart changes these)
        price = ""
        for sel in ["._30jeq3", "._1_WHN1", ".Nx9bqj"]:
            el = soup.select_one(sel)
            if el:
                price = el.get_text(strip=True)
                break

        title = ""
        for sel in [".s1Q9rs", "._4rR01T", ".WKTcLC"]:
            el = soup.select_one(sel)
            if el:
                title = el.get_text(strip=True)
                break

        link = f"https://www.flipkart.com/search?q={isbn}"
        return {"title": title, "price": price, "link": link}
    except Exception:
        return {"link": f"https://www.flipkart.com/search?q={isbn}"}

# ── routes ─────────────────────────────────────────────────────────────────

@app.route("/lookup", methods=["POST"])
def lookup():
    """Lookup a single ISBN from all sources."""
    body = request.get_json()
    isbn = body.get("isbn", "").strip()
    if not isbn:
        return jsonify({"error": "No ISBN provided"}), 400

    google = fetch_google_books(isbn)
    ol     = fetch_open_library(isbn)
    amazon = fetch_amazon(isbn)
    flipkart = fetch_flipkart(isbn)

    return jsonify({
        "isbn":     isbn,
        "google":   google,
        "openlibrary": ol,
        "amazon":   amazon,
        "flipkart": flipkart,
    })

@app.route("/bulk-excel", methods=["POST"])
def bulk_excel():
    """Accept a list of ISBNs, fetch all data, return an Excel file."""
    body = request.get_json()
    isbns = body.get("isbns", [])
    if not isbns:
        return jsonify({"error": "No ISBNs provided"}), 400

    results = []
    for isbn in isbns:
        isbn = isbn.strip()
        if not isbn:
            continue
        google   = fetch_google_books(isbn)
        ol       = fetch_open_library(isbn)
        amazon   = fetch_amazon(isbn)
        flipkart = fetch_flipkart(isbn)
        results.append({
            "isbn": isbn,
            "google": google,
            "openlibrary": ol,
            "amazon": amazon,
            "flipkart": flipkart,
        })
        time.sleep(0.3)  # be polite to servers

    # ── Build Excel ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Sheet 1: Combined view
    ws = wb.active
    ws.title = "All Data"
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1D3557")
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "ISBN",
        "Title (GB)", "Author (GB)", "Publisher (GB)", "Year (GB)", "Pages (GB)", "Lang (GB)",
        "Title (OL)", "Author (OL)", "Publisher (OL)", "Year (OL)", "Pages (OL)",
        "Amazon Price", "Amazon Link",
        "Flipkart Price", "Flipkart Link",
    ]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center

    for r in results:
        g  = r["google"]
        ol = r["openlibrary"]
        am = r["amazon"]
        fk = r["flipkart"]
        ws.append([
            r["isbn"],
            g.get("title",""), g.get("authors",""), g.get("publisher",""),
            g.get("year",""), g.get("pages",""), g.get("language",""),
            ol.get("title",""), ol.get("authors",""), ol.get("publisher",""),
            ol.get("year",""), ol.get("pages",""),
            am.get("price",""), am.get("link",""),
            fk.get("price",""), fk.get("link",""),
        ])

    # Auto column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Sheet 2: Amazon detail
    ws2 = wb.create_sheet("Amazon")
    ws2.append(["ISBN", "Title found", "Price", "Search link"])
    for r in results:
        am = r["amazon"]
        ws2.append([r["isbn"], am.get("title",""), am.get("price",""), am.get("link","")])

    # Sheet 3: Flipkart detail
    ws3 = wb.create_sheet("Flipkart")
    ws3.append(["ISBN", "Title found", "Price", "Search link"])
    for r in results:
        fk = r["flipkart"]
        ws3.append([r["isbn"], fk.get("title",""), fk.get("price",""), fk.get("link","")])

    # Sheet 4: Store links
    ws4 = wb.create_sheet("Store Links")
    ws4.append(["ISBN", "Amazon link", "Flipkart link"])
    for r in results:
        ws4.append([r["isbn"], r["amazon"].get("link",""), r["flipkart"].get("link","")])

    # Send file back to browser
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="books.xlsx",
    )

@app.route("/health")
def health():
    return "ok"

if __name__ == "__main__":
    app.run(debug=True)
